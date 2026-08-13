import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import xlrd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .import_rules import EXPECTED_HEADERS
from .schemas import ImportIssue

MAX_FILE_SIZE = 5 * 1024 * 1024
MAX_DATA_ROWS = 500
ORDER_ID_PATTERN = re.compile(r"^[A-Za-z0-9_-]{1,64}$")


@dataclass
class ParsedOrder:
    order_id: str
    customer_name: str
    amount: Decimal
    order_date: date


class ExcelValidationError(Exception):
    def __init__(self, message: str, errors: list[ImportIssue] | None = None) -> None:
        self.message = message
        self.errors = errors or []
        super().__init__(message)


def _string(value: Any) -> str:
    return value.strip() if isinstance(value, str) else str(value).strip()


def _read_xlsx(content: bytes) -> list[list[Any]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    except Exception as error:
        raise ExcelValidationError("无法读取 XLSX 文件，请确认文件未损坏或未加密") from error

    sheet = workbook.active
    rows: list[list[Any]] = []
    for row in sheet.iter_rows():
        values: list[Any] = []
        for cell in row:
            if cell.data_type == "f":
                raise ExcelValidationError(
                    "不允许使用公式单元格",
                    [ImportIssue(row=cell.row, field=None, message="请将公式转换为固定值后再上传")],
                )
            values.append(cell.value)
        rows.append(values)
    workbook.close()
    return rows


def _read_xls(content: bytes) -> list[list[Any]]:
    try:
        workbook = xlrd.open_workbook(file_contents=content)
    except Exception as error:
        raise ExcelValidationError("无法读取 XLS 文件，请确认文件未损坏或未加密") from error

    sheet = workbook.sheet_by_index(0)
    rows: list[list[Any]] = []
    for row_index in range(sheet.nrows):
        values: list[Any] = []
        for column_index in range(sheet.ncols):
            cell = sheet.cell(row_index, column_index)
            if cell.ctype == xlrd.XL_CELL_DATE:
                values.append(xlrd.xldate_as_datetime(cell.value, workbook.datemode).date())
            else:
                values.append(cell.value)
        rows.append(values)
    return rows


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def validate_order_values(
    order_id_value: Any,
    customer_name_value: Any,
    amount_value: Any,
    order_date_value: Any,
    *,
    row: int | None = None,
) -> ParsedOrder:
    """Validate one order using the same rules as an Excel data row."""
    order_id = _string(order_id_value) if order_id_value is not None else ""
    customer_name = _string(customer_name_value) if customer_name_value is not None else ""
    errors: list[ImportIssue] = []

    if not ORDER_ID_PATTERN.fullmatch(order_id):
        errors.append(ImportIssue(row=row, field="订单ID", message="必填，只允许字母、数字、下划线和连字符，最多 64 位"))
    if not customer_name or len(customer_name) > 100:
        errors.append(ImportIssue(row=row, field="客户名称", message="必填，且最多 100 个字符"))

    amount: Decimal | None = None
    try:
        if isinstance(amount_value, bool) or amount_value in (None, ""):
            raise InvalidOperation
        amount = Decimal(str(amount_value))
        if amount <= 0 or amount.as_tuple().exponent < -2 or amount > Decimal("9999999999.99"):
            raise InvalidOperation
    except (InvalidOperation, ValueError):
        errors.append(ImportIssue(row=row, field="订单金额", message="必填，必须是大于 0 且最多两位小数的数字"))

    order_date = _parse_date(order_date_value)
    if order_date is None:
        errors.append(ImportIssue(row=row, field="下单日期", message="必填，格式必须为 YYYY-MM-DD"))

    if errors:
        raise ExcelValidationError("订单数据校验失败，请按错误提示修改后重试", errors)
    assert amount is not None and order_date is not None
    return ParsedOrder(order_id, customer_name, amount, order_date)


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _without_blank_trailing_columns(rows: list[list[Any]]) -> list[list[Any]]:
    required_columns = len(EXPECTED_HEADERS)
    for row_number, row in enumerate(rows, start=1):
        for column_number, value in enumerate(row[required_columns:], start=required_columns + 1):
            if not _is_blank(value):
                column = get_column_letter(column_number)
                raise ExcelValidationError(
                    "Excel 数据校验失败，请按错误提示修改后重新上传",
                    [
                        ImportIssue(
                            row=row_number,
                            field=f"{column}列",
                            message=f"第 {column} 列存在未识别数据，请删除第 {column} 列及其后的额外数据",
                        )
                    ],
                )
    return [row[:required_columns] for row in rows]


def parse_orders(filename: str, content: bytes) -> tuple[list[ParsedOrder], int]:
    suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if suffix not in {"xlsx", "xls"}:
        raise ExcelValidationError("只支持 .xlsx 或 .xls 格式的 Excel 文件")
    if not content:
        raise ExcelValidationError("上传文件为空")
    if len(content) > MAX_FILE_SIZE:
        raise ExcelValidationError("文件不能超过 5 MB")

    rows = _read_xlsx(content) if suffix == "xlsx" else _read_xls(content)
    if not rows:
        raise ExcelValidationError("Excel 文件没有内容")
    rows = _without_blank_trailing_columns(rows)

    headers = tuple(_string(value) if value is not None else "" for value in rows[0])
    if headers != EXPECTED_HEADERS:
        raise ExcelValidationError(
            "表头不符合模板要求",
            [
                ImportIssue(
                    row=1,
                    field="表头",
                    message=f"必须完全按顺序填写：{'、'.join(EXPECTED_HEADERS)}；当前为：{'、'.join(headers) or '空'}",
                )
            ],
        )

    data_rows = [row for row in rows[1:] if any(value not in (None, "") for value in row)]
    if not data_rows:
        raise ExcelValidationError("Excel 文件没有可导入的数据行")
    if len(data_rows) > MAX_DATA_ROWS:
        raise ExcelValidationError(f"单次最多导入 {MAX_DATA_ROWS} 行")

    orders: list[ParsedOrder] = []
    errors: list[ImportIssue] = []
    seen_order_ids: set[str] = set()
    for offset, values in enumerate(data_rows, start=2):
        if len(values) != len(EXPECTED_HEADERS):
            errors.append(ImportIssue(row=offset, field=None, message="字段数量必须与模板完全一致"))
            continue

        order_id = _string(values[0]) if values[0] is not None else ""
        is_duplicate_order_id = order_id in seen_order_ids
        if not ORDER_ID_PATTERN.fullmatch(order_id):
            errors.append(ImportIssue(row=offset, field="订单ID", message="必填，只允许字母、数字、下划线和连字符，最多 64 位"))
        elif is_duplicate_order_id:
            errors.append(ImportIssue(row=offset, field="订单ID", message="文件内订单ID重复"))
        else:
            seen_order_ids.add(order_id)
        try:
            order = validate_order_values(*values, row=offset)
            if not ORDER_ID_PATTERN.fullmatch(order_id) or is_duplicate_order_id:
                continue
            orders.append(order)
        except ExcelValidationError as error:
            errors.extend(error.errors)

    if errors:
        raise ExcelValidationError("Excel 数据校验失败，请按错误提示修改后重新上传", errors)
    return orders, len(data_rows)
