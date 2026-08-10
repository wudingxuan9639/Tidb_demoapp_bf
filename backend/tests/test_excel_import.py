import io
import unittest

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.excel_import import ExcelValidationError, parse_orders
from app.import_rules import EXPECTED_HEADERS


def workbook_bytes(headers: tuple[str, ...], row: list[object]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def workbook_with_trailing_column_bytes(trailing_value: object | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(EXPECTED_HEADERS)
    sheet.append(["ORD-10001", "上海示例客户", 1288.50, "2026-08-09"])

    # WPS/Excel can retain a blank, formatted column in the worksheet range.
    sheet["E1"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    sheet["E2"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    if trailing_value is not None:
        sheet["E2"] = trailing_value

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


class ParseOrdersTest(unittest.TestCase):
    def test_valid_workbook_maps_chinese_headers_to_order_fields(self) -> None:
        content = workbook_bytes(
            EXPECTED_HEADERS, ["ORD-10001", "上海示例客户", 1288.50, "2026-08-09"]
        )

        orders, row_count = parse_orders("orders.xlsx", content)

        self.assertEqual(row_count, 1)
        self.assertEqual(orders[0].order_id, "ORD-10001")
        self.assertEqual(orders[0].customer_name, "上海示例客户")
        self.assertEqual(str(orders[0].amount), "1288.5")
        self.assertEqual(orders[0].order_date.isoformat(), "2026-08-09")

    def test_wrong_header_returns_clear_validation_error(self) -> None:
        content = workbook_bytes(
            ("订单编号", "客户名称", "订单金额", "下单日期"),
            ["ORD-10001", "上海示例客户", 1288.50, "2026-08-09"],
        )

        with self.assertRaisesRegex(ExcelValidationError, "表头不符合模板要求") as error:
            parse_orders("orders.xlsx", content)

        self.assertEqual(error.exception.errors[0].row, 1)

    def test_invalid_amount_returns_row_and_field(self) -> None:
        content = workbook_bytes(
            EXPECTED_HEADERS, ["ORD-10001", "上海示例客户", "免费", "2026-08-09"]
        )

        with self.assertRaisesRegex(ExcelValidationError, "Excel 数据校验失败") as error:
            parse_orders("orders.xlsx", content)

        self.assertEqual(error.exception.errors[0].row, 2)
        self.assertEqual(error.exception.errors[0].field, "订单金额")

    def test_blank_formatted_trailing_columns_are_ignored(self) -> None:
        orders, row_count = parse_orders("orders.xlsx", workbook_with_trailing_column_bytes())

        self.assertEqual(row_count, 1)
        self.assertEqual(orders[0].order_id, "ORD-10001")

    def test_trailing_column_value_returns_location_specific_error(self) -> None:
        with self.assertRaisesRegex(ExcelValidationError, "Excel 数据校验失败") as error:
            parse_orders("orders.xlsx", workbook_with_trailing_column_bytes("备注"))

        issue = error.exception.errors[0]
        self.assertEqual(issue.row, 2)
        self.assertEqual(issue.field, "E列")
        self.assertIn("第 E 列存在未识别数据", issue.message)
